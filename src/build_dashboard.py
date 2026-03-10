"""
build_dashboard.py
Builds an Excel dashboard with 3 sheets + charts using openpyxl.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
import os

DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "credit_data_clean.csv")
OUT_PATH  = os.path.join(os.path.dirname(__file__), "..", "output", "dashboard.xlsx")

df = pd.read_csv(DATA_PATH)

# ── Styles ─────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
ACCENT_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="DEEAF1")
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL  = PatternFill("solid", fgColor="FFEB9C")
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(bold=True, size=14, color="1F4E79")
BOLD_FONT    = Font(bold=True, size=11)
NORMAL_FONT  = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

thin = Side(style="thin", color="BDD7EE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — Credit Utilization by Risk Segment
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Credit Utilization"
ws1.sheet_view.showGridLines = False

# Title
ws1["B2"] = "Credit Utilization Analysis by Risk Segment"
ws1["B2"].font   = TITLE_FONT
ws1["B2"].alignment = LEFT

# Aggregated table
util_summary = df.groupby("risk_label").agg(
    Customers=("customer_id", "count"),
    Avg_Utilization=("credit_utilization", "mean"),
    Low_Util_pct=("credit_utilization", lambda x: (x < 0.3).mean() * 100),
    Med_Util_pct=("credit_utilization", lambda x: ((x >= 0.3) & (x < 0.6)).mean() * 100),
    High_Util_pct=("credit_utilization", lambda x: (x >= 0.6).mean() * 100),
).round(2).reset_index()
util_summary = util_summary.set_index("risk_label").loc[["Low", "Medium", "High"]].reset_index()
util_summary.columns = ["Risk Label", "Customers", "Avg Utilization", "Low (<30%)", "Medium (30-60%)", "High (>60%)"]

headers = list(util_summary.columns)
start_row = 4

for col_idx, header in enumerate(headers, start=2):
    cell = ws1.cell(row=start_row, column=col_idx, value=header)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER

risk_fills = {"Low": GREEN_FILL, "Medium": YELLOW_FILL, "High": RED_FILL}
for row_idx, row_data in util_summary.iterrows():
    excel_row = start_row + 1 + row_idx
    fill = risk_fills.get(row_data["Risk Label"], WHITE_FILL)
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws1.cell(row=excel_row, column=col_idx, value=val)
        cell.fill      = fill
        cell.alignment = CENTER
        cell.border    = BORDER
        cell.font      = NORMAL_FONT

# KPI boxes
kpi_row = 10
ws1[f"B{kpi_row}"] = "Average Credit Utilization"
ws1[f"C{kpi_row}"] = f"{df['credit_utilization'].mean()*100:.1f}%"
ws1[f"E{kpi_row}"] = "High Utilization Customers (>60%)"
ws1[f"F{kpi_row}"] = f"{(df['credit_utilization'] > 0.6).sum()}"
ws1[f"H{kpi_row}"] = "Total Portfolio"
ws1[f"I{kpi_row}"] = f"${df['loan_amount'].sum():,.0f}"

for cell_ref in [f"B{kpi_row}", f"E{kpi_row}", f"H{kpi_row}"]:
    ws1[cell_ref].font = BOLD_FONT
    ws1[cell_ref].fill = ACCENT_FILL
    ws1[cell_ref].font = Font(bold=True, color="FFFFFF")
    ws1[cell_ref].alignment = CENTER
    ws1[cell_ref].border = BORDER
for cell_ref in [f"C{kpi_row}", f"F{kpi_row}", f"I{kpi_row}"]:
    ws1[cell_ref].font = BOLD_FONT
    ws1[cell_ref].alignment = CENTER
    ws1[cell_ref].border = BORDER

# Bar chart — Avg Utilization by Risk
chart1 = BarChart()
chart1.type          = "col"
chart1.title         = "Average Credit Utilization by Risk Label"
chart1.y_axis.title  = "Avg Utilization"
chart1.x_axis.title  = "Risk Label"
chart1.height        = 12
chart1.width         = 18

data_ref = Reference(ws1, min_col=3, max_col=3, min_row=start_row, max_row=start_row + 3)
cats_ref = Reference(ws1, min_col=2, max_col=2, min_row=start_row + 1, max_row=start_row + 3)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.solidFill = "2E75B6"
ws1.add_chart(chart1, "B13")

# Column widths
for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
    ws1.column_dimensions[col].width = 22
ws1.row_dimensions[2].height = 25

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — Repayment Trends
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Repayment Trends")
ws2.sheet_view.showGridLines = False

ws2["B2"] = "Repayment Trends & Default Analysis"
ws2["B2"].font      = TITLE_FONT
ws2["B2"].alignment = LEFT

# Repayment bucket analysis
df["repayment_bucket"] = pd.cut(
    df["repayment_history"], bins=[0, 8, 16, 20, 24],
    labels=["Poor (0-8)", "Fair (9-16)", "Good (17-20)", "Excellent (21-24)"],
    include_lowest=True,
)
rep_summary = df.groupby("repayment_bucket", observed=True).agg(
    Customers=("customer_id", "count"),
    Avg_Credit_Score=("credit_score", "mean"),
    Avg_Defaults=("num_defaults", "mean"),
    High_Risk_Pct=("risk_label", lambda x: (x == "High").mean() * 100),
).round(2).reset_index()
rep_summary.columns = ["Repayment Band", "Customers", "Avg Credit Score", "Avg Defaults", "High Risk %"]

start_row2 = 4
for col_idx, header in enumerate(rep_summary.columns, start=2):
    cell = ws2.cell(row=start_row2, column=col_idx, value=header)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = BORDER

for row_idx, row_data in rep_summary.iterrows():
    excel_row = start_row2 + 1 + row_idx
    fill = [RED_FILL, YELLOW_FILL, ALT_FILL, GREEN_FILL][row_idx]
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws2.cell(row=excel_row, column=col_idx, value=val)
        cell.fill = fill; cell.alignment = CENTER
        cell.border = BORDER; cell.font = NORMAL_FONT

# Default count by num_defaults value
def_dist = df.groupby("num_defaults").agg(
    Customers=("customer_id", "count")
).reset_index()
def_dist.columns = ["Number of Defaults", "Customers"]

start_row2b = 12
for col_idx, header in enumerate(def_dist.columns, start=2):
    cell = ws2.cell(row=start_row2b, column=col_idx, value=header)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = BORDER

for row_idx, row_data in def_dist.iterrows():
    excel_row = start_row2b + 1 + row_idx
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws2.cell(row=excel_row, column=col_idx, value=val)
        cell.alignment = CENTER; cell.border = BORDER
        cell.fill = ALT_FILL if row_idx % 2 else WHITE_FILL
        cell.font = NORMAL_FONT

# Line chart — Avg defaults per repayment band
chart2 = BarChart()
chart2.type        = "col"
chart2.title       = "Average Defaults by Repayment Quality"
chart2.y_axis.title = "Avg Defaults"
chart2.height      = 12; chart2.width = 18
data_ref2 = Reference(ws2, min_col=5, max_col=5, min_row=start_row2, max_row=start_row2 + 4)
cats_ref2 = Reference(ws2, min_col=2, max_col=2, min_row=start_row2 + 1, max_row=start_row2 + 4)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
chart2.series[0].graphicalProperties.solidFill = "E74C3C"
ws2.add_chart(chart2, "G4")

for col in ["B", "C", "D", "E", "F"]:
    ws2.column_dimensions[col].width = 22

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — Customer Risk Scores
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Customer Risk Scores")
ws3.sheet_view.showGridLines = False

ws3["B2"] = "Customer Risk Score Summary"
ws3["B2"].font = TITLE_FONT; ws3["B2"].alignment = LEFT

# Full risk summary by risk label
risk_summary = df.groupby("risk_label").agg(
    Customers=("customer_id", "count"),
    Avg_Credit_Score=("credit_score", "mean"),
    Avg_Income=("income", "mean"),
    Avg_DTI=("debt_to_income_ratio", "mean"),
    Avg_Utilization=("credit_utilization", "mean"),
    Avg_Defaults=("num_defaults", "mean"),
    Total_Loan=("loan_amount", "sum"),
).round(2).reset_index()
risk_summary = risk_summary.set_index("risk_label").loc[["Low", "Medium", "High"]].reset_index()
risk_summary.columns = ["Risk Label", "Customers", "Avg Credit Score", "Avg Income",
                         "Avg DTI", "Avg Utilization", "Avg Defaults", "Total Loan Exposure"]

start_row3 = 4
for col_idx, header in enumerate(risk_summary.columns, start=2):
    cell = ws3.cell(row=start_row3, column=col_idx, value=header)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = BORDER

for row_idx, row_data in risk_summary.iterrows():
    excel_row = start_row3 + 1 + row_idx
    fill = [GREEN_FILL, YELLOW_FILL, RED_FILL][row_idx]
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws3.cell(row=excel_row, column=col_idx, value=val)
        cell.fill = fill; cell.alignment = CENTER
        cell.border = BORDER; cell.font = NORMAL_FONT

# KPIs
kpi_items = [
    ("B9", "Total Customers"),   ("C9", f"{len(df):,}"),
    ("E9", "High Risk Count"),   ("F9", f"{(df['risk_label']=='High').sum():,}"),
    ("H9", "Portfolio at Risk"), ("I9", f"${df[df['risk_label']=='High']['loan_amount'].sum():,.0f}"),
    ("B10", "Avg Credit Score"), ("C10", f"{df['credit_score'].mean():.0f}"),
    ("E10", "Avg Income"),       ("F10", f"${df['income'].mean():,.0f}"),
    ("H10", "Avg DTI Ratio"),    ("I10", f"{df['debt_to_income_ratio'].mean():.3f}"),
]
for cell_ref, val in kpi_items:
    ws3[cell_ref] = val
    ws3[cell_ref].alignment = CENTER
    ws3[cell_ref].border = BORDER
    if cell_ref[1] == "9" or cell_ref[1] == "1":
        ws3[cell_ref].fill = ACCENT_FILL
        ws3[cell_ref].font = Font(bold=True, color="FFFFFF")

# Bar chart — Total Loan Exposure by Risk
chart3 = BarChart()
chart3.type        = "col"
chart3.title       = "Total Loan Exposure by Risk Tier"
chart3.y_axis.title = "Loan Amount ($)"
chart3.height      = 12; chart3.width = 18
data_ref3 = Reference(ws3, min_col=9, max_col=9, min_row=start_row3, max_row=start_row3 + 3)
cats_ref3 = Reference(ws3, min_col=2, max_col=2, min_row=start_row3 + 1, max_row=start_row3 + 3)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
chart3.series[0].graphicalProperties.solidFill = "1F4E79"
ws3.add_chart(chart3, "B13")

for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
    ws3.column_dimensions[col].width = 22

# ── Save ───────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
wb.save(OUT_PATH)
print(f"Dashboard saved → {os.path.abspath(OUT_PATH)}")
